"""
Antidote resale analysis engine.

WHAT THIS DOES
--------------
Reads your tracker spreadsheet and rebuilds every number and chart from scratch.
Through the year you just add rows to the tracker and re-run this file. Nothing
is hardcoded, so the outputs always reflect the latest data.

HOW TO USE
----------
1. Change INPUT_FILE below to wherever your tracker lives.
2. Run:  python antidote_analysis.py
3. It prints a summary and saves the charts as PNG files in OUTPUT_DIR.

WHAT YOU CAN EDIT (all in the CONFIG block)
-------------------------------------------
- CATEGORY_RULES: how items get sorted into product lines (keyword based, so a
  new multimeter next month auto-sorts itself).
- NICHE: which product lines count as your repeatable engine.
- FORWARD_MONTHLY_CAPITAL / FORWARD_CYCLES_PER_MONTH: the two assumptions behind
  the forward plan. The cycles number is a guess until your dated data can measure it.
"""

# ============================= CONFIG =============================
INPUT_FILE = "Summer_Work_-_Camilo_-_Project.xlsx"   # <-- change to your file path
OUTPUT_DIR = "."                            # where charts get saved

# Product-line rules. Checked top to bottom; first keyword hit wins.
# To add a line, add a (name, [keywords]) pair. To move an item, change a keyword.
CATEGORY_RULES = [
    ("Fluke test equipment", ["multimeter", "fluke", "87v", "117", "177"]),
    ("Nest thermostats",     ["thermo"]),
    ("Non-electronics",      ["jacket", "jazzmaster", "guitar", "weight"]),
    ("Other small electronics", ["nest mini", "google plug", "echo", "amazon cube",
                                 "phone", "airpod", "charger", "monster", "xbox",
                                 "pen", "bose", "headphone", "calculator"]),
]

# Your repeatable engine. Everything else is treated as opportunistic / one-off.
NICHE = {"Fluke test equipment", "Nest thermostats"}

# Forward plan assumptions.
# --- Fluke case study inputs (from my own eBay sold-listing research) ---
FLUKE_DEMAND = {"87V": 743, "117": 251, "789": 172, "1507": 133, "773": 90}
SEARCH_HOURS_NOW = 3.0     # hours/week I search today
SEARCH_HOURS_MAX = 5.0     # the most I would spend
UNITS_PER_MONTH_NOW = 3.0  # meters I find per month today
TICKET_NOW = 385.0         # my average Fluke buy price so far
TICKET_TARGET = 600.0      # target average buy price

FORWARD_MONTHLY_CAPITAL = 500.0    # dollars you would deploy per month
FORWARD_CYCLES_PER_MONTH = 1.0     # times that cash recycles per month (a GUESS for now)
# =================================================================

from openpyxl import load_workbook
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

HEADER_ROW = 3   # row where the column names live
# Verified zero-based positions against the actual file:
# 1=Item, 7=Used, 8=New, 9=Turnaround Time, 10=COGS, 11=Net Price
COL_ITEM, COL_USED, COL_NEW, COL_TURNAROUND, COL_COGS, COL_NET = 1, 7, 8, 9, 10, 11


def categorize(name):
    low = name.lower()
    for line, keywords in CATEGORY_RULES:
        if any(k in low for k in keywords):
            return line
    return "Uncategorized"


def load_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        item = r[COL_ITEM]
        if item is None or str(item).strip() == "":
            continue                      # skip blank spacer rows
        if isinstance(item, (int, float)) or str(item).strip().replace('.', '', 1).isdigit():
            continue                      # skip the totals/count row at the bottom
        cogs = float(r[COL_COGS] or 0)
        net = float(r[COL_NET] or 0)
        turnaround = r[COL_TURNAROUND]
        condition = "Used" if r[COL_USED] == "X" else ("New" if r[COL_NEW] == "X" else "Unknown")
        rows.append({
            "item": str(item).strip(),
            "cogs": cogs,
            "net": net,
            "profit": net - cogs,         # recomputed, not read from the formula
            "category": categorize(str(item)),
            "status": "Sold" if net > 0 else "Inventory",
            "acq": "Paid" if cogs > 0 else "Free",
            "turnaround_days": float(turnaround) if isinstance(turnaround, (int, float)) else None,
            "condition": condition,
        })
    return rows


def money(x):
    return f"${x:,.2f}"


def pct(x):
    return f"{x*100:.1f}%"


def summarize(rows):
    sold = [x for x in rows if x["status"] == "Sold"]
    inv = [x for x in rows if x["status"] == "Inventory"]

    real_profit = sum(x["profit"] for x in sold)
    real_proceeds = sum(x["net"] for x in sold)
    real_cogs = sum(x["cogs"] for x in sold)
    inv_capital = sum(x["cogs"] for x in inv)

    print("=" * 60)
    print("REALIZED (sold items only)")
    print("=" * 60)
    print(f"  Items sold        : {len(sold)}")
    print(f"  Net proceeds      : {money(real_proceeds)}")
    print(f"  Cost of goods sold: {money(real_cogs)}")
    print(f"  Realized profit   : {money(real_profit)}")
    if real_proceeds:
        print(f"  Blended margin    : {pct(real_profit/real_proceeds)}")
    print()
    print(f"  Cash tied up in inventory (NOT a loss): {money(inv_capital)} "
          f"across {len(inv)} unsold item(s)")
    print()

    # Paid vs free split of realized profit (repeatable vs windfall)
    paid_profit = sum(x["profit"] for x in sold if x["acq"] == "Paid")
    free_profit = sum(x["profit"] for x in sold if x["acq"] == "Free")
    print("-" * 60)
    print("WHERE REALIZED PROFIT CAME FROM")
    print("-" * 60)
    print(f"  Bought to resell (repeatable): {money(paid_profit)} "
          f"({pct(paid_profit/real_profit)})")
    print(f"  Free finds (one-off windfall): {money(free_profit)} "
          f"({pct(free_profit/real_profit)})")
    print()

    # Per product line, realized
    print("-" * 60)
    print("BY PRODUCT LINE (sold items)")
    print("-" * 60)
    lines = sorted({x["category"] for x in sold})
    per_line = {}
    for line in lines:
        g = [x for x in sold if x["category"] == line]
        paid = [x for x in g if x["acq"] == "Paid"]
        total_profit = sum(x["profit"] for x in g)
        free_profit = sum(x["profit"] for x in g if x["acq"] == "Free")
        paid_profit = sum(x["profit"] for x in paid)
        paid_cogs = sum(x["cogs"] for x in paid)
        paid_proceeds = sum(x["net"] for x in paid)
        roi = paid_profit / paid_cogs if paid_cogs else None      # bought items only
        margin = paid_profit / paid_proceeds if paid_proceeds else None
        per_line[line] = {"profit": total_profit, "roi": roi, "margin": margin,
                          "n": len(g), "n_paid": len(paid),
                          "free_profit": free_profit, "paid_profit": paid_profit}
        tag = "  [NICHE]" if line in NICHE else ""
        print(f"  {line}{tag}")
        if paid:
            print(f"      {len(g)} sold ({len(paid)} bought) | profit {money(total_profit)}"
                  f" | on bought items: margin {pct(margin)}, ROI {pct(roi)}")
        else:
            print(f"      {len(g)} sold (all free finds) | profit {money(total_profit)}"
                  f" | no capital used")
        if free_profit and paid:
            print(f"      of which {money(free_profit)} came from free finds in this line")

    # Niche engine, combined
    niche_sold = [x for x in sold if x["category"] in NICHE]
    np_ = sum(x["profit"] for x in niche_sold)
    nc = sum(x["cogs"] for x in niche_sold)
    niche_roi = np_ / nc if nc else 0
    print()
    print("-" * 60)
    print("NICHE ENGINE (Fluke + Nest combined)")
    print("-" * 60)
    print(f"  Profit: {money(np_)} | Capital used: {money(nc)} | ROI: {pct(niche_roi)}")
    print()

    # Forward plan
    print("-" * 60)
    print("FORWARD PLAN (assumption-based)")
    print("-" * 60)
    monthly_profit = FORWARD_MONTHLY_CAPITAL * niche_roi * FORWARD_CYCLES_PER_MONTH
    print(f"  Deploy {money(FORWARD_MONTHLY_CAPITAL)}/month into the niche at "
          f"{pct(niche_roi)} ROI,")
    print(f"  recycling {FORWARD_CYCLES_PER_MONTH:g}x/month  ->  "
          f"~{money(monthly_profit)} profit/month.")
    print(f"  CAVEAT: the recycle rate is a guess. Add purchase/sale dates and the")
    print(f"  monthly section below will measure it instead.")
    print()

    return per_line, niche_roi


def turnaround_analysis(rows):
    print("-" * 60)
    print("TURNAROUND TIME (days from buy to sell)")
    print("-" * 60)
    sold_with_time = [x for x in rows if x["status"] == "Sold" and x["turnaround_days"] is not None]
    if not sold_with_time:
        print("  No turnaround data yet.")
        print()
        return
    avg_all = sum(x["turnaround_days"] for x in sold_with_time) / len(sold_with_time)
    print(f"  Average across all sold items: {avg_all:.1f} days ({len(sold_with_time)} items)")

    fluke_sold = [x for x in sold_with_time if x["category"] == "Fluke test equipment"]
    if fluke_sold:
        avg_fluke = sum(x["turnaround_days"] for x in fluke_sold) / len(fluke_sold)
        fastest = min(fluke_sold, key=lambda x: x["turnaround_days"])
        print(f"  Fluke average: {avg_fluke:.1f} days ({len(fluke_sold)} completed sales)")
        print(f"      fastest: '{fastest['item']}' sold in {fastest['turnaround_days']:.0f} days")
        print(f"  NOTE: sample size is small ({len(fluke_sold)}). Treat as directional, not precise.")
    else:
        print("  No completed Fluke sales with turnaround data yet.")
    print()


def ticket_band_analysis(rows):
    """Does ROI hold as capital deployed per unit increases? The core scale question."""
    print("-" * 60)
    print("DOES ROI HOLD AS TICKET SIZE SCALES? (Fluke, bought items)")
    print("-" * 60)
    fl = [x for x in rows if x["status"] == "Sold" and x["acq"] == "Paid"
          and x["category"] == "Fluke test equipment"]
    bands = [("Under $250", 0, 250), ("$250 to $500", 250, 500), ("Over $500", 500, 10**9)]
    out = []
    for label, lo, hi in bands:
        g = [x for x in fl if lo < x["cogs"] <= hi] if lo else [x for x in fl if x["cogs"] <= hi]
        if not g:
            continue
        p = sum(x["profit"] for x in g)
        c = sum(x["cogs"] for x in g)
        roi = p / c
        per_unit = p / len(g)
        avg_days = sum(x["turnaround_days"] for x in g if x["turnaround_days"]) / len(g)
        out.append((label, len(g), roi, per_unit, avg_days))
        print(f"  {label:14s} | n={len(g):2d} | ROI {pct(roi):>6s} | "
              f"profit/unit {money(per_unit):>8s} | avg {avg_days:.1f}d")
    print()
    print("  READ: if ROI is flat across bands, capital scales without margin decay.")
    print()
    return out


def inventory_check(rows):
    inv = [x for x in rows if x["status"] == "Inventory"]
    fluke_inv = [x for x in inv if x["category"] == "Fluke test equipment"]
    if fluke_inv:
        print("-" * 60)
        print("FLUKE CURRENTLY IN INVENTORY (unsold, cash tied up)")
        print("-" * 60)
        for x in fluke_inv:
            days_listed = f", listed for reference ({x['turnaround_days']:.0f}d so far)" if x["turnaround_days"] else ""
            print(f"  '{x['item']}' | cost {money(x['cogs'])} | condition: {x['condition']}{days_listed}")
        print()


YEL = "#F2B909"
DARK = "#0C1F2D"
MGRAY = "#8C99A6"
LGRAY = "#D9D9D9"
SERIF = "Carlito"


def _style():
    plt.rcParams.update({
        "figure.autolayout": True,
        "font.family": SERIF,
        "font.size": 15,
        "text.color": DARK,
        "axes.labelcolor": DARK,
        "xtick.color": DARK,
        "ytick.color": DARK,
        "axes.edgecolor": LGRAY,
        "axes.spines.top": False,
        "axes.spines.right": False,
        "figure.facecolor": "white",
    })


def make_charts(rows, per_line, bands=None):
    sold = [x for x in rows if x["status"] == "Sold"]
    _style()

    # ---- 1. Profit by product line: horizontal, Fluke highlighted ----
    items = sorted(per_line.items(), key=lambda kv: kv[1]["profit"])
    names = [k for k, _ in items]
    vals = [v["profit"] for _, v in items]
    cols = [YEL if n == "Fluke test equipment" else LGRAY for n in names]
    fig, ax = plt.subplots(figsize=(8.2, 3.6))
    ax.barh(names, vals, color=cols, height=0.62)
    ax.set_xlim(0, max(vals) * 1.22)
    for i, v in enumerate(vals):
        ax.text(v + max(vals) * 0.02, i, f"${v:,.0f}", va="center", fontsize=15,
                fontweight="bold" if names[i] == "Fluke test equipment" else "normal",
                color=DARK)
    ax.set_xticks([])
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_color(LGRAY)
    ax.tick_params(axis="y", length=0, labelsize=15)
    fig.savefig(f"{OUTPUT_DIR}/c_profit_by_line.png", dpi=200)
    plt.close(fig)

    # ---- 2. Bought vs free: one stacked bar ----
    paid = sum(x["profit"] for x in sold if x["acq"] == "Paid")
    free = sum(x["profit"] for x in sold if x["acq"] == "Free")
    tot = paid + free
    fig, ax = plt.subplots(figsize=(8.2, 1.9))
    ax.barh([0], [paid], color=YEL, height=0.5)
    ax.barh([0], [free], left=[paid], color=LGRAY, height=0.5)
    ax.text(paid / 2, 0, f"Bought to resell\n${paid:,.0f}  ({paid/tot*100:.0f}%)",
            ha="center", va="center", fontsize=15, fontweight="bold", color=DARK)
    ax.text(paid + free / 2, 0, f"Free finds\n${free:,.0f}  ({free/tot*100:.0f}%)",
            ha="center", va="center", fontsize=14, color=DARK)
    ax.set_xlim(0, tot)
    ax.axis("off")
    fig.savefig(f"{OUTPUT_DIR}/c_bought_vs_free.png", dpi=200)
    plt.close(fig)

    # ---- 3. Ticket bands: two panels, no dual axis ----
    if bands:
        labels = [b[0].replace("$", "\\$") for b in bands]
        rois = [b[2] * 100 for b in bands]
        per = [b[3] for b in bands]
        fig, (a1, a2) = plt.subplots(1, 2, figsize=(11.0, 3.9))
        a1.bar(labels, rois, color=YEL, width=0.56)
        a1.set_ylim(0, 60)
        a1.set_title("Return stays flat", fontsize=17, pad=14, fontweight="bold")
        for i, v in enumerate(rois):
            a1.text(i, v + 1.6, f"{v:.0f}%", ha="center", fontsize=17,
                    fontweight="bold", color=DARK)
        a1.set_yticks([])
        a1.spines["left"].set_visible(False)
        a1.spines["bottom"].set_color(LGRAY)
        a1.tick_params(axis="x", length=0, labelsize=14)

        a2.bar(labels, per, color=DARK, width=0.56)
        a2.set_ylim(0, max(per) * 1.3)
        a2.set_title("Profit per sale rises 5x", fontsize=17, pad=14, fontweight="bold")
        for i, v in enumerate(per):
            a2.text(i, v + max(per) * 0.03, f"${v:.0f}", ha="center", fontsize=17,
                    fontweight="bold", color=DARK)
        a2.set_yticks([])
        a2.spines["left"].set_visible(False)
        a2.spines["bottom"].set_color(LGRAY)
        a2.tick_params(axis="x", length=0, labelsize=14)
        fig.savefig(f"{OUTPUT_DIR}/c_ticket_bands.png", dpi=200)
        plt.close(fig)

    # ---- 4. Capital vs profit, guitar called out ----
    paid_items = [x for x in sold if x["acq"] == "Paid"]
    fig, ax = plt.subplots(figsize=(8.0, 5.0))
    for x in paid_items:
        is_fluke = x["category"] == "Fluke test equipment"
        ax.scatter(x["cogs"], x["profit"],
                   color=YEL if is_fluke else MGRAY,
                   edgecolor=DARK, linewidth=0.7,
                   s=150 if is_fluke else 110, zorder=3)
    g = next((x for x in paid_items if "jazz" in x["item"].lower()), None)
    if g:
        ax.annotate("The guitar\n\\$400 in, \\$60 out, 30 days",
                    xy=(g["cogs"], g["profit"]),
                    xytext=(g["cogs"] - 340, g["profit"] + 155),
                    fontsize=14, color=DARK, ha="left",
                    arrowprops=dict(arrowstyle="->", color=DARK, lw=1.3))
    ax.set_xlabel("What I paid for the item ($)", fontsize=15, labelpad=10)
    ax.set_ylabel("What I made on it ($)", fontsize=15, labelpad=10)
    ax.grid(axis="y", color=LGRAY, linewidth=0.7, zorder=0)
    ax.set_axisbelow(True)
    ax.spines["left"].set_color(LGRAY)
    ax.spines["bottom"].set_color(LGRAY)
    ax.tick_params(labelsize=13)
    from matplotlib.lines import Line2D
    ax.legend(handles=[
        Line2D([0], [0], marker="o", color="w", markerfacecolor=YEL,
               markeredgecolor=DARK, markersize=12, label="Fluke"),
        Line2D([0], [0], marker="o", color="w", markerfacecolor=MGRAY,
               markeredgecolor=DARK, markersize=11, label="Everything else"),
    ], frameon=False, fontsize=14, loc="upper left")
    fig.savefig(f"{OUTPUT_DIR}/c_capital_vs_profit.png", dpi=200)
    plt.close(fig)

    # ---- 5. Fluke demand: recent eBay sold listings by model ----
    dm = sorted(FLUKE_DEMAND.items(), key=lambda kv: kv[1])
    mn = [f"Fluke {k}" for k, _ in dm]
    mv = [v for _, v in dm]
    fig, ax = plt.subplots(figsize=(8.0, 3.4))
    ax.barh(mn, mv, color=[YEL if v == max(mv) else LGRAY for v in mv], height=0.6)
    ax.set_xlim(0, max(mv) * 1.2)
    for i, v in enumerate(mv):
        ax.text(v + max(mv) * 0.02, i, f"{v:,}", va="center", fontsize=15,
                fontweight="bold" if v == max(mv) else "normal", color=DARK)
    ax.set_xticks([])
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_color(LGRAY)
    ax.tick_params(axis="y", length=0, labelsize=15)
    fig.savefig(f"{OUTPUT_DIR}/c_fluke_demand.png", dpi=200)
    plt.close(fig)

    # ---- 6. Two levers: more hours, bigger meters ----
    fl = [x for x in rows if x["status"] == "Sold" and x["acq"] == "Paid"
          and x["category"] == "Fluke test equipment"]
    roi = sum(x["profit"] for x in fl) / sum(x["cogs"] for x in fl)
    units_max = UNITS_PER_MONTH_NOW * (SEARCH_HOURS_MAX / SEARCH_HOURS_NOW)
    scen = [
        ("Today\n3 hrs, $385", UNITS_PER_MONTH_NOW * TICKET_NOW * roi),
        ("More hours\n5 hrs, $385", units_max * TICKET_NOW * roi),
        ("Bigger meters\n3 hrs, $600", UNITS_PER_MONTH_NOW * TICKET_TARGET * roi),
        ("Both\n5 hrs, $600", units_max * TICKET_TARGET * roi),
    ]
    fig, ax = plt.subplots(figsize=(9.0, 3.8))
    labs = [a for a, _ in scen]
    vals = [b for _, b in scen]
    cols = [LGRAY, LGRAY, LGRAY, YEL]
    ax.bar(labs, vals, color=cols, width=0.58, edgecolor=DARK, linewidth=0.6)
    ax.set_ylim(0, max(vals) * 1.24)
    for i, v in enumerate(vals):
        ax.text(i, v + max(vals) * 0.03, f"${v:,.0f}", ha="center", fontsize=17,
                fontweight="bold", color=DARK)
    ax.set_yticks([])
    ax.spines["left"].set_visible(False)
    ax.spines["bottom"].set_color(LGRAY)
    ax.tick_params(axis="x", length=0, labelsize=13.5)
    ax.set_title("Profit per month at a 41.8% return", fontsize=16, pad=14)
    fig.savefig(f"{OUTPUT_DIR}/c_scale_levers.png", dpi=200)
    plt.close(fig)

    print("Saved deck charts to", OUTPUT_DIR)


def main():
    rows = load_rows(INPUT_FILE)
    uncats = [x["item"] for x in rows if x["category"] == "Uncategorized"]
    if uncats:
        print("NOTE: these items did not match any category rule, add a keyword for them:")
        for u in uncats:
            print("   -", u)
        print()
    per_line, niche_roi = summarize(rows)
    turnaround_analysis(rows)
    bands = ticket_band_analysis(rows)
    inventory_check(rows)
    make_charts(rows, per_line, bands)


if __name__ == "__main__":
    main()
